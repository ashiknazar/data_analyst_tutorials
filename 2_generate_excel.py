import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment
from datetime import datetime

df=pd.read_csv("sample_data.csv")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS
excel_filename = f"styled_data_{timestamp}.xlsx"

df.to_excel(excel_filename,index=False,engine="openpyxl")

wb=load_workbook(excel_filename)
ws=wb.active

# Define styles
header_font = Font(bold=True, color="FFFFFF")  # White text
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
alignment = Alignment(horizontal="center")

# Apply styles to header
for cell in ws[1]:  
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter (e.g., 'A', 'B', etc.)
    
    for cell in col:
        try:  # Handle cases where cell value is None
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    
    ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

# Save the styled Excel file
wb.save(excel_filename)